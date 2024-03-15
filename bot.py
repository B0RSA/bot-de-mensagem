import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook

os.environ['WDM_LOG_LEVEL'] = '0'

def iniciar_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    return webdriver.Chrome(options=options)

def enviar_mensagem_whatsapp(contato, mensagem, driver):
    driver.get("https://web.whatsapp.com/")
    driver.implicitly_wait(10) # tempo de espera para a página carregar(aumentar se precisar)
    input("Faça o login no WhatsApp Web e pressione Enter após a conclusão...")

    try:
        caixa_de_pesquisa = driver.find_element_by_xpath("//div[@contenteditable='true'][@data-tab='3']")
        caixa_de_pesquisa.send_keys(contato)
        caixa_de_pesquisa.send_keys(Keys.ENTER)
        caixa_de_mensagem = driver.find_element_by_xpath("//div[@contenteditable='true'][@data-tab='1']")
        caixa_de_mensagem.send_keys(mensagem)
        caixa_de_mensagem.send_keys(Keys.ENTER)
    except TimeoutException as e:
        print("Tempo limite excedido ao aguardar o elemento.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
    finally:
        driver.quit()

def ler_numeros_telefone_planilha(nome_arquivo, nome_planilha, coluna_numero):
    numeros = []
    try:
        planilha = load_workbook(nome_arquivo)
        folha = planilha[nome_planilha]
        coluna_numero = coluna_numero.upper()
        coluna = folha[coluna_numero]
        print(f"Lendo números da coluna {coluna_numero}")
        for cell in coluna[1:]:
            numero = str(cell.value)
            print(f"Numero encontrado: {numero}")
            if numero:
                numeros.append(numero)
    except FileNotFoundError:
        print("Arquivo não encontrado. Verifique se o nome e o caminho do arquivo estão corretos.")
    except KeyError:
        print("A planilha especificada não existe no arquivo. Verifique o nome da planilha.")
    except Exception as e:
        print(f"Erro ao ler a planilha: {e}")
    return numeros

def listar_planilhas(nome_arquivo):
    try:
        planilha = load_workbook(nome_arquivo, read_only=True)
        return planilha.sheetnames
    except FileNotFoundError:
        print("Arquivo não encontrado. Verifique se o nome e o caminho do arquivo estão corretos.")
    except Exception as e:
        print(f"Erro ao ler o arquivo Excel: {e}")
        return []

if __name__ == "__main__":
    arquivo_excel = input("Digite o nome do arquivo (inclua a extensão .xlsx): ")
    planilhas = listar_planilhas(arquivo_excel)
    if planilhas:
        print("Planilhas disponíveis no arquivo:")
        for planilha in planilhas:
            print(planilha)
    else:
        print("Não foi possível obter a lista de planilhas.")
    nome_planilha = input("Digite o nome da planilha (Ex: Planilha1): ")
    coluna_numero = input("Digite a letra da coluna em que se encontram os números: ")
    mensagem = input("Digite a mensagem a ser enviada: ")

    driver = iniciar_driver()
    numeros_telefone = ler_numeros_telefone_planilha(arquivo_excel, nome_planilha, coluna_numero)
    for numero in numeros_telefone:
        enviar_mensagem_whatsapp(numero, mensagem, driver)
